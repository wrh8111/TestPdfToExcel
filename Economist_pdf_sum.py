import sys

import camelot.io as camelot
import pandas as pd
import openpyxl
import re
import os


class Economist_pdf_sum:

    def __init__(self, pdf_file):
        self.__temp_file = 'temp.xlsx'
        self.pdf_file = pdf_file

        if os.path.exists(self.__temp_file) == False:
            tables = camelot.read_pdf(self.pdf_file, pages='1-end', flavor='stream')
            res = pd.concat([table.df for table in tables], axis=0, ignore_index=True)
            res.to_excel(self.__temp_file, encoding='utf8')

    def close(self):
        self.__clear_temp_xlsx()

    #如果temp.xlsx文件存在，则删除
    def __clear_temp_xlsx(self):
        if os.path.exists(self.__temp_file):
            os.remove((self.__temp_file))

    #从pdf文件提取表格文本数据到temp.xlsx文件
    # def __create_temp_xlsx(self):
    #     if os.path.exists(self.__temp_file) == False:
    #         tables = camelot.read_pdf(self.pdf_file, pages='1-end', flavor='stream')
    #         res = pd.concat([table.df for table in tables], axis=0, ignore_index=True)
    #         res.to_excel(self.__temp_file, encoding='utf8')


    #统计PDF文件中‘姓名’字符出现的个数（通俗认为，多少个‘姓名’，就有多少人）
    #具体实现流程：
    #1.提取PDF文件的表格到temp.xlsx文件中
    #2.统计temp.xlsx文件中‘姓名’字符出现的个数
    def get_psum_from_file(self):
        sum = 0
        if os.path.exists(self.__temp_file):
            wb = openpyxl.load_workbook(self.__temp_file)
            ws = wb.active
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    if re.match('姓名', str(ws.cell(row, col).value).strip()) != None:
                        sum += 1
            wb.close()
        return sum

    #统计pdf文件中每场的人数（如：第一场多少人，第二场多少人）
    def get_psum_from_NFS(self):
        sum = 0
        NFC = ''
        NFC_count = []
        if os.path.exists(self.__temp_file):
            wb = openpyxl.load_workbook(self.__temp_file)
            ws = wb.active
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    this_cell = str(ws.cell(row, col).value).strip()

                    if len(re.findall(r'第\d+场', this_cell)) == 1:
                        NFC_cell = re.findall(r'第\d+场', this_cell)[0]
                        if NFC_cell != NFC:
                            NFC_count.append(sum)
                            NFC = NFC_cell
                            sum = 0

                    if re.match('姓名', this_cell) != None:
                        sum += 1
            wb.close()
            NFC_count.append(sum)
        return NFC_count[1:]

# 生成统计报表Excel文件
def get_report(director):
    dir_str = director
    files = os.listdir(dir_str)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = '考场文件名称'
    for col in range(1, 9):
        ws.cell(1, col + 1).value = '第' + str(col) + '场'
    ws.cell(1, 10).value = '考场总人数'
    li_row = []#用于装所有统计数据
    for file in files:
        # print(dir_str+file)
        if '.pdf' in file:
            aa = Economist_pdf_sum(dir_str + file)
            # print(file[:-4],aa.get_psum_from_NFS(),aa.get_psum_from_file())
            li_row.append([file[:-4], aa.get_psum_from_NFS(), aa.get_psum_from_file()])
            aa.close()
    i = 2 #作为数据行的指针
    cc = 2#作为场次列的指针
    for row in li_row:
        ws.cell(i, 1).value = row[0]
        for c in row[1]:
            ws.cell(i, cc).value = c
            cc += 1
        cc = 2
        ws.cell(i, 10).value = row[2]
        i += 1
    wb.save('Report.xlsx')

if __name__ == '__main__':
    # aa = Economist_pdf_sum('202110_001_1350201001.pdf')
    # # print(aa.get_psum_from_file())
    # print(aa.get_psum_from_NFS(),aa.get_psum_from_file())
    # aa.close()
    # aa = Economist_pdf_sum('202110_001_1350201008.pdf')
    # # print(aa.get_psum_from_file())
    # print(aa.get_psum_from_NFS(),aa.get_psum_from_file())
    # aa.close()
    # dir_str = 'D:\\Work\\2021年数据\\经济\\经济机考\\2021年各地市经济机考考场数据\\厦门\\[1350201]厦门技师学院_20211017124910\\全部座次表\\'
    # dir_str = os.getcwd()+'\\'
    if sys.argv[1] == None:
        get_report(os.getcwd()+'\\')
    else:
        get_report(sys.argv[1] + '\\')