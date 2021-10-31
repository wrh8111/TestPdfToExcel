# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('经济参缺表.xlsx', data_only=True)
ws = wb.active
# //////////////////////////////////////
# exam_name = '基础知识'
exam_name = '专业知识与实务'
row1 = 36
# /////////////////////////////////////

row2 = row1 + 1
row3 = row2 + 1
print(
    '10月31日上午第六场经济%s，厦门考区应考%s人， 实考%s人 ，缺考%s人，参考率%.2f%%。'%(exam_name,str(ws.cell(row1, 4).value), str(ws.cell(row1, 5).value), str(ws.cell(row1, 6).value),
                                                          float(ws.cell(row1, 7).value)*100))
print(
    '10月31日上午两场总计，厦门考区应考%s人， 实考%s人 ，缺考%s人，参考率%.2f%%。' % (str(ws.cell(row2, 4).value), str(ws.cell(row2, 5).value), str(ws.cell(row2, 6).value),
    float(ws.cell(row2, 7).value*100)))
print(
    '10月30、31日六场总计，厦门考区应考%s人， 实考%s人 ，缺考%s人，参考率%.2f%%。' % (str(ws.cell(row3, 4).value), str(ws.cell(row3, 5).value), str(ws.cell(row3, 6).value),
    float(ws.cell(row3, 7).value*100)))